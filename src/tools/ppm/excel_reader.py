from dataclasses import dataclass
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import load_workbook


@dataclass
class Resumen:
    activo: Any = 1  # default 1 para compatibilidad si viene vacío
    folio_ppm: str = ""
    nombre_proyecto: str = ""
    objetivo: str = ""
    horas_internas: int = 0
    horas_externas: int = 0
    horas_totales: int = 0
    costo_total: float = 0.0
    fecha_inicio: Any = None
    fecha_fin_liberacion: Any = None
    fecha_fin_garantia: Any = None
    avance_planeado: float = 0.0
    avance_real: float = 0.0
    direccion_area: str = ""
    lider_cliente: str = ""
    ern: str = ""
    le: str = ""
    ppm: str = ""
    descripcion_estatus: str = ""


@dataclass
class GanttRow:
    folio_ppm: str
    actividad: str
    responsable: str
    fecha_inicio: Any
    fecha_fin: Any
    avance: Any
    color: str
    estatus_etapa: Optional[str] = None


@dataclass
class RiesgoRow:
    folio_ppm: str
    riesgo: str
    responsable: str
    mitigacion: str
    fecha_materializacion: Any


def _sheet_to_dicts(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() for h in rows[0]]
    out = []
    for r in rows[1:]:
        if all(v is None for v in r):
            continue
        d = {headers[i]: r[i] for i in range(len(headers))}
        out.append(d)
    return out


def _filter_keys(d: Dict[str, Any], cls) -> Dict[str, Any]:
    allowed = set(getattr(cls, "__dataclass_fields__", {}).keys())
    return {k: v for k, v in d.items() if k in allowed}


def read_all_projects(excel_path: str) -> List[Tuple[Resumen, List[GanttRow], List[RiesgoRow]]]:
    wb = load_workbook(excel_path, data_only=True)

    resumen_rows = _sheet_to_dicts(wb["RESUMEN"])
    gantt_rows = _sheet_to_dicts(wb["GANTT"])
    riesgos_rows = _sheet_to_dicts(wb["RIESGOS"])

    # index rápido por folio
    gantt_by = {}
    for g in gantt_rows:
        folio = str(g.get("folio_ppm", "")).strip()
        if not folio:
            continue
        g2 = _filter_keys(g, GanttRow)
        gantt_by.setdefault(folio, []).append(GanttRow(**g2))

    riesgos_by = {}
    for r in riesgos_rows:
        folio = str(r.get("folio_ppm", "")).strip()
        if not folio:
            continue
        r2 = _filter_keys(r, RiesgoRow)
        riesgos_by.setdefault(folio, []).append(RiesgoRow(**r2))

    projects = []
    for r in resumen_rows:
        r2 = _filter_keys(r, Resumen)
        resumen = Resumen(**r2)
        projects.append((resumen, gantt_by.get(resumen.folio_ppm, []), riesgos_by.get(resumen.folio_ppm, [])))

    return projects
